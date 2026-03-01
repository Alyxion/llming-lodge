"""XLSX text extraction using openpyxl, preserving hyperlinks."""
import io
from pathlib import Path
from typing import Union


def extract_xlsx(source: Union[Path, bytes]) -> str:
    from openpyxl import load_workbook

    fp = io.BytesIO(source) if isinstance(source, bytes) else source
    # read_only=False needed to access cell hyperlinks
    wb = load_workbook(fp, data_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row is None or ws.max_row < 1:
            continue

        # Build markdown table
        lines = [f"**Sheet: {sheet_name}**", ""]
        first_row = True
        for row in ws.iter_rows():
            cells = []
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                link = cell.hyperlink
                if link and link.target and val:
                    val = f"[{val}]({link.target})"
                elif link and link.target:
                    val = link.target
                cells.append(val)
            line = "| " + " | ".join(cells) + " |"
            lines.append(line)
            if first_row:
                lines.append("| " + " | ".join(["---"] * len(cells)) + " |")
                first_row = False
        sheets.append("\n".join(lines))
    wb.close()
    return "\n\n".join(sheets) if sheets else "[No data found in spreadsheet]"
