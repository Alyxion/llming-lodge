"""Server-side table export to XLSX and CSV formats.

Uses openpyxl for XLSX (already a project dependency) and csv for CSV.
"""

import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)


def _normalize_spec(spec: dict) -> tuple[list[str], list[list[Any]]]:
    """Normalize a table spec into (columns, rows).

    Handles both flat format ``{columns: [...], rows: [[...]]}`` and
    object-row format ``{columns: [{key, label}], rows: [{key: val}]}``.

    Returns:
        Tuple of (column_labels, row_data) where row_data is a list of
        flat lists (one value per column).
    """
    raw_cols = spec.get("columns", [])
    raw_rows = spec.get("rows", [])

    # Build column labels and keys
    col_labels: list[str] = []
    col_keys: list[str] = []
    for c in raw_cols:
        if isinstance(c, dict):
            col_labels.append(str(c.get("label", c.get("key", ""))))
            col_keys.append(str(c.get("key", c.get("label", ""))))
        else:
            col_labels.append(str(c))
            col_keys.append(str(c))

    # Normalize rows
    rows: list[list[Any]] = []
    for row in raw_rows:
        if isinstance(row, dict):
            rows.append([row.get(k, "") for k in col_keys])
        elif isinstance(row, (list, tuple)):
            rows.append(list(row))
        else:
            rows.append([row])

    return col_labels, rows


def export_xlsx(spec: dict) -> bytes:
    """Generate XLSX from a table spec.

    Args:
        spec: Table spec with ``columns`` and ``rows``.

    Returns:
        XLSX file content as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    columns, rows = _normalize_spec(spec)
    title = spec.get("title", "Sheet1")

    wb = Workbook()
    ws = wb.active
    # Excel forbids these chars in sheet names: \ / * ? : [ ]
    safe_title = (title or "Sheet1")
    for ch in r'\/*?:[]':
        safe_title = safe_title.replace(ch, "_")
    ws.title = safe_title[:31]  # Excel sheet name limit

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1D459F", end_color="1D459F", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    if columns:
        for col_idx, label in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # Write data rows
    start_row = 2 if columns else 1
    cell_font = Font(name="Calibri", size=11)
    for row_idx, row_data in enumerate(rows, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_coerce_value(value))
            cell.font = cell_font
            cell.border = thin_border

    # Auto-width columns (approximate)
    for col_idx in range(1, (len(columns) or 1) + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max(max_len + 2, 8), 50)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_csv(spec: dict) -> bytes:
    """Generate CSV from a table spec.

    Args:
        spec: Table spec with ``columns`` and ``rows``.

    Returns:
        UTF-8 encoded CSV content as bytes.
    """
    columns, rows = _normalize_spec(spec)

    buf = io.StringIO()
    writer = csv.writer(buf)
    if columns:
        writer.writerow(columns)
    for row in rows:
        writer.writerow([_coerce_value(v) for v in row])

    return buf.getvalue().encode("utf-8")


def _coerce_value(value: Any) -> Any:
    """Coerce a cell value to a native Python type for Excel/CSV.

    Attempts numeric conversion for string values that look like numbers.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    s = str(value).strip()
    if not s:
        return ""
    # Try int
    try:
        return int(s)
    except (ValueError, OverflowError):
        pass
    # Try float
    try:
        return float(s)
    except (ValueError, OverflowError):
        pass
    return s
